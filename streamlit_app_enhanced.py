import streamlit as st
import requests
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
from openai import OpenAI
import pandas as pd
import os
import re
import io
from io import BytesIO
from dotenv import load_dotenv

# Import our custom modules
from edgar_enhanced import (
    lookup_cik, get_ticker_from_cik, get_accessions, 
    get_ex99_1_links, find_guidance_paragraphs
)
from transcript_provider import get_transcript_for_quarter
from guidance_extractor import (
    extract_guidance, extract_transcript_guidance, 
    process_guidance_table
)
import importlib
import duplicate_handler
importlib.reload(duplicate_handler)
from duplicate_handler import (
    detect_duplicates, highlight_duplicates, reset_duplicate_state
)
from supabase_store import (
    upload_user_document, get_uploaded_documents, 
    download_document, delete_document, get_supabase_config
)
from document_tagger import extract_document_metadata, extract_text_from_file, validate_and_confirm_metadata

# Load environment variables
load_dotenv()

def format_percent(val):
    """Format a value as a percentage with consistent decimal places"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return f"{val:.1f}%"
    return val

def format_dollar(val):
    """Format a value as a dollar amount with consistent decimal places"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if abs(val) >= 100:
            return f"${val:.0f}"
        elif abs(val) >= 10:
            return f"${val:.1f}"
        else:
            return f"${val:.2f}"
    return val

def is_cik_format(s):
    return s.isdigit() and len(s) == 10

# ---- Streamlit UI Setup ----

st.set_page_config(page_title="Enhanced SEC 8-K & Transcript Guidance Extractor", layout="centered")
st.title("Enhanced SEC 8-K & Transcript Guidance Extractor")
st.markdown("**Extract Guidance from SEC filings and Earnings Call Transcripts with AI.**")

# API Key Management with Password Protection
st.sidebar.header("🔐 API Configuration")

# Password for using hosted API key
app_password = st.sidebar.text_input("Enter app password to use hosted OpenAI key:", type="password", key="app_password")

# Option to use own API key
use_own_key = st.sidebar.checkbox("Use my own OpenAI API key instead")

if use_own_key:
    user_openai_key = st.sidebar.text_input("Enter your OpenAI API key:", type="password", key="user_openai_key")
    if user_openai_key:
        openai_key = user_openai_key
    else:
        st.error("Please enter your OpenAI API key to continue")
        st.stop()
else:
    # Check password for hosted key
    try:
        hosted_key = st.secrets["OPENAI_API_KEY"]
        correct_password = st.secrets.get("APP_PASSWORD", "guidance2025")  # Default password if not set
        
        if app_password == correct_password:
            openai_key = hosted_key
            st.sidebar.success("✅ Using hosted OpenAI key")
        else:
            if app_password:  # Only show error if they tried entering a password
                st.sidebar.error("❌ Incorrect password")
            st.error("Please enter the correct app password or use your own OpenAI API key")
            st.stop()
    except KeyError:
        st.error("Hosted OpenAI key not configured. Please use your own API key.")
        st.stop()


# Create main tabs
main_tab1, main_tab2 = st.tabs(["🎯 Guidance Extraction", "📁 Document Management"])

with main_tab1:
    st.header("Extract Guidance from Multiple Sources")
    
    # Input section
    ticker_or_cik = st.text_input(
        "Enter Stock Ticker or CIK (e.g., MSFT or 0000789019)",
        "MSFT",
        help="Enter either a stock ticker (e.g., MSFT) or a 10-digit CIK code (e.g., 0000789019)"
    )

    openai_models = {
        "GPT-4o Mini": "gpt-4o-mini",
        "GPT-4 Turbo": "gpt-4-turbo-preview",
        "GPT-4": "gpt-4",
        "GPT-3.5 Turbo": "gpt-3.5-turbo"
    }
    selected_model = st.selectbox(
        "Select OpenAI Model",
        list(openai_models.keys()),
        index=0
    )

    # Data source selection
    st.subheader("Data Sources")
    col1, col2, col3 = st.columns(3)
    with col1:
        extract_sec = st.checkbox("SEC 8-K Filings", value=True)
    with col2:
        extract_transcripts = st.checkbox("Earnings Transcripts", value=True)
    with col3:
        extract_uploaded = st.checkbox("Uploaded Documents", value=True)

    # Time period selection
    st.subheader("Time Period")
    year_input = st.text_input("How many years back to search for filings? (Leave blank for most recent only)", "")
    quarter_input = st.text_input("OR enter specific quarter (e.g., 2Q25, Q4FY24)", "")

with main_tab2:
    st.header("Document Management")
    
    # Document Management Section
    doc_tab1, doc_tab2 = st.tabs(["📤 Upload Documents", "📋 View Uploaded Documents"])

    with doc_tab1:
        st.write("Upload earnings documents for a specific ticker")
        
        # Ticker input first
        col1, col2, col3, col4 = st.columns([2, 2, 2, 1])
        with col1:
            upload_ticker = st.text_input("Ticker Symbol", placeholder="e.g., AAPL", help="All uploaded files will be tagged with this ticker").upper()
        with col2:
            upload_year = st.number_input("Year (optional)", min_value=2020, max_value=2030, value=None, help="Leave blank to auto-detect from documents")
        with col3:
            upload_quarter = st.selectbox("Quarter (optional)", ["", "Q1", "Q2", "Q3", "Q4"], help="Leave blank to auto-detect from documents")
        with col4:
            if st.button("🗑️ Clear", help="Clear all form fields"):
                st.session_state.clear()
                st.rerun()
        
        
        # File uploader (always visible)
        if upload_ticker:
            year_text = str(upload_year) if upload_year else "auto-detect"
            quarter_text = upload_quarter if upload_quarter else "auto-detect"
            uploaded_files = st.file_uploader(
                f"Choose files for {upload_ticker} ({year_text}, {quarter_text})", 
                type=['pdf', 'txt', 'html', 'docx'],
                key="file_uploader",
                accept_multiple_files=True,
                help=f"Upload multiple earnings documents for {upload_ticker}. Each file will be analyzed individually for year/quarter."
            )
        else:
            uploaded_files = st.file_uploader(
                "Choose earnings documents", 
                type=['pdf', 'txt', 'html', 'docx'],
                key="file_uploader",
                accept_multiple_files=True,
                help="Upload multiple earnings documents. Enter ticker symbol above first."
            )
        
        # Show file validation info
        if uploaded_files:
            st.info(f"📁 Selected {len(uploaded_files)} files: {', '.join([f.name for f in uploaded_files])}")
            
            # Basic file validation
            total_size = sum(f.size for f in uploaded_files if hasattr(f, 'size'))
            if total_size > 50 * 1024 * 1024:  # 50MB limit
                st.warning("⚠️ Total file size is large. Processing may take longer.")
            
            # Check for duplicate filenames
            filenames = [f.name for f in uploaded_files]
            if len(filenames) != len(set(filenames)):
                st.error("❌ Duplicate filenames detected. Please rename files to have unique names.")
                st.stop()
        
        # Upload button (always visible when files are selected)
        if uploaded_files:
            st.write(f"📁 **{len(uploaded_files)} files selected** - Ready to analyze")
            
            # Initialize session state for batch processing
            if "batch_documents" not in st.session_state:
                st.session_state.batch_documents = []
            if "batch_analyzed" not in st.session_state:
                st.session_state.batch_analyzed = False
            
            # Step 1: Analyze documents
            if not st.session_state.batch_analyzed:
                if st.button("🔍 Analyze Documents", type="primary", use_container_width=True):
                    # Validate ticker input first
                    if not upload_ticker or len(upload_ticker.strip()) < 1:
                        st.error("❌ Please enter a ticker symbol before analyzing documents")
                        st.stop()
                    
                    # Initialize OpenAI client for document type extraction
                    if not openai_key:
                        st.error("OpenAI API key is required for document analysis")
                        st.stop()
                    
                    client = OpenAI(api_key=openai_key)
                    
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    # Clear previous analysis results
                    st.session_state.batch_documents = []
                    
                    for i, uploaded_file in enumerate(uploaded_files):
                        status_text.text(f"Processing {uploaded_file.name}...")
                        progress_bar.progress((i + 1) / len(uploaded_files))
                        
                        try:
                            # Read file data
                            file_data = uploaded_file.read()
                            content_type = uploaded_file.type or "application/octet-stream"
                            
                            # Validate file data type
                            if not isinstance(file_data, bytes):
                                raise Exception(f"File read returned {type(file_data)} instead of bytes: {repr(file_data)}")
                            
                            # Extract text content with better error handling
                            file_extension = uploaded_file.name.split('.')[-1].lower() if '.' in uploaded_file.name else 'unknown'
                            
                            # Validate file size
                            if len(file_data) > 10 * 1024 * 1024:  # 10MB per file
                                raise Exception(f"File too large ({len(file_data)/1024/1024:.1f}MB). Maximum size is 10MB.")
                            
                            # Validate file extension
                            allowed_extensions = ['pdf', 'txt', 'html', 'docx']
                            if file_extension not in allowed_extensions:
                                raise Exception(f"Unsupported file type: {file_extension}. Allowed: {', '.join(allowed_extensions)}")
                            
                            text_content = extract_text_from_file(file_data, file_extension)
                            
                            # Validate extracted text
                            if not text_content or len(text_content.strip()) < 50:
                                raise Exception("Could not extract sufficient text content from file")
                            
                            # Extract metadata using AI (year, quarter, document type)
                            metadata = None
                            if text_content and len(text_content.strip()) > 100:
                                try:
                                    metadata = extract_document_metadata(text_content, uploaded_file.name, client)
                                except Exception as ai_error:
                                    st.warning(f"AI analysis failed for {uploaded_file.name}: {str(ai_error)}")
                                    metadata = None
                            
                            # Use AI-extracted year/quarter if available, otherwise use user input as fallback
                            final_year = metadata.get('year') if metadata and metadata.get('year') else upload_year
                            final_quarter = metadata.get('quarter') if metadata and metadata.get('quarter') else upload_quarter
                            
                            # Ensure we have valid year/quarter (required for upload)
                            if not final_year:
                                raise Exception("Could not determine year from document content and no fallback year provided")
                            if not final_quarter:
                                raise Exception("Could not determine quarter from document content and no fallback quarter provided")
                            
                            # Store analyzed document for review
                            doc_info = {
                                'filename': uploaded_file.name,
                                'file_data': file_data,
                                'content_type': content_type,
                                'ticker': upload_ticker,
                                'detected_year': metadata.get('year') if metadata else None,
                                'detected_quarter': metadata.get('quarter') if metadata else None,
                                'detected_doc_type': metadata.get('document_type') if metadata else None,
                                'fallback_year': upload_year,
                                'fallback_quarter': upload_quarter,
                                'final_year': final_year,
                                'final_quarter': final_quarter,
                                'text_content': text_content[:500] + '...' if len(text_content) > 500 else text_content
                            }
                            st.session_state.batch_documents.append(doc_info)
                            
                        except Exception as e:
                            st.error(f"❌ Failed to process {uploaded_file.name}: {str(e)}")
                    
                    # Mark analysis as complete
                    st.session_state.batch_analyzed = True
                    status_text.text("Analysis complete!")
                    progress_bar.progress(1.0)
                    
                    st.success(f"✅ Analyzed {len(st.session_state.batch_documents)} documents!")
                    st.rerun()
            
            # Step 2: Review and approve metadata
            else:
                st.write("📋 **Review Analyzed Documents**")
                
                if st.session_state.batch_documents:
                    # Show analyzed documents for review
                    for i, doc in enumerate(st.session_state.batch_documents):
                        with st.expander(f"📄 {doc['filename']}", expanded=True):
                            col1, col2, col3 = st.columns([2, 2, 2])
                            
                            with col1:
                                st.write("**Detected Metadata:**")
                                st.write(f"• Year: {doc['detected_year'] or 'Not detected'}")
                                st.write(f"• Quarter: {doc['detected_quarter'] or 'Not detected'}")
                                st.write(f"• Document Type: {doc['detected_doc_type'] or 'Not detected'}")
                            
                            with col2:
                                st.write("**Edit Metadata:**")
                                # Editable year field
                                edited_year = st.number_input(
                                    "Year", 
                                    min_value=2020, 
                                    max_value=2030, 
                                    value=doc['final_year'],
                                    key=f"edit_year_{i}"
                                )
                                # Editable quarter field
                                quarter_options = ["Q1", "Q2", "Q3", "Q4"]
                                current_quarter_index = quarter_options.index(doc['final_quarter']) if doc['final_quarter'] in quarter_options else 0
                                edited_quarter = st.selectbox(
                                    "Quarter",
                                    quarter_options,
                                    index=current_quarter_index,
                                    key=f"edit_quarter_{i}"
                                )
                                
                                # Update the document in session state
                                st.session_state.batch_documents[i]['final_year'] = edited_year
                                st.session_state.batch_documents[i]['final_quarter'] = edited_quarter
                            
                            with col3:
                                st.write("**Final Values (for upload):**")
                                st.write(f"• Ticker: {doc['ticker']}")
                                st.write(f"• Year: {edited_year}")
                                st.write(f"• Quarter: {edited_quarter}")
                            
                            # Show text preview
                            st.write("**Text Preview:**")
                            st.text_area("Document Preview", doc['text_content'], height=100, disabled=True, key=f"preview_{i}", label_visibility="hidden")
                    
                    # Upload buttons
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("✅ Upload All Documents", type="primary", use_container_width=True):
                            # Perform actual upload
                            progress_bar = st.progress(0)
                            status_text = st.empty()
                            uploaded_count = 0
                            failed_count = 0
                            
                            for i, doc in enumerate(st.session_state.batch_documents):
                                status_text.text(f"Uploading {doc['filename']}...")
                                progress_bar.progress((i + 1) / len(st.session_state.batch_documents))
                                
                                try:
                                    result = upload_user_document(
                                        ticker=doc['ticker'],
                                        year=doc['final_year'],
                                        quarter=doc['final_quarter'],
                                        filename=doc['filename'],
                                        file_data=doc['file_data'],
                                        content_type=doc['content_type']
                                    )
                                    uploaded_count += 1
                                    st.success(f"✅ Uploaded: {doc['filename']}")
                                except Exception as upload_error:
                                    failed_count += 1
                                    error_msg = f"❌ Failed to upload {doc['filename']}: {str(upload_error)}"
                                    st.error(error_msg)
                                    
                                    # Store detailed error for persistence
                                    if 'upload_errors' not in st.session_state:
                                        st.session_state.upload_errors = []
                                    st.session_state.upload_errors.append({
                                        'filename': doc['filename'],
                                        'error': str(upload_error),
                                        'timestamp': str(pd.Timestamp.now())
                                    })
                            
                            # Show final results
                            status_text.text("Upload complete!")
                            if uploaded_count > 0:
                                st.success(f"🎉 Successfully uploaded {uploaded_count} documents!")
                                st.balloons()
                            if failed_count > 0:
                                st.error(f"❌ Failed to upload {failed_count} documents")
                            
                            # Store error details in session state for persistence
                            if failed_count > 0:
                                if 'upload_errors' not in st.session_state:
                                    st.session_state.upload_errors = []
                                # Keep only the last 10 errors to avoid memory issues
                                st.session_state.upload_errors = st.session_state.upload_errors[-9:]
                            
                            # Don't clear session state immediately - let user see results
                            # Clear session state
                            # st.session_state.batch_documents = []
                            # st.session_state.batch_analyzed = False
                            # st.rerun()
                    
                    with col2:
                        if st.button("🔄 Re-analyze", use_container_width=True):
                            st.session_state.batch_analyzed = False
                            st.session_state.batch_documents = []
                            def reset_duplicate_state():
                                """Reset duplicate resolution state"""
                                keys_to_remove = [k for k in st.session_state.keys() if k.startswith('dup_')]
                                for key in keys_to_remove:
                                    del st.session_state[key]
                                
                                # Also clear stored duplicates and combined data
                                if 'stored_duplicates' in st.session_state:
                                    del st.session_state['stored_duplicates']
                                if 'stored_combined' in st.session_state:
                                    del st.session_state['stored_combined']
                                if 'duplicate_selections' in st.session_state:
                                    del st.session_state['duplicate_selections']
                                if 'duplicates_resolved' in st.session_state:
                                    del st.session_state['duplicates_resolved']
                                if 'cleaned_df' in st.session_state:
                                    del st.session_state['cleaned_df']
                            reset_duplicate_state()
                            st.rerun()
                else:
                    st.info("No documents analyzed yet.")
            
        
        # Clear session state button
        if st.button("🔄 Start Over", help="Clear all files and start fresh"):
            st.session_state.batch_documents = []
            st.session_state.batch_analyzed = False
            if 'upload_errors' in st.session_state:
                st.session_state.upload_errors = []
            st.rerun()
        

    with doc_tab2:
        st.write("View and filter uploaded documents")
        
        # Action buttons
        col1, col2 = st.columns([1, 4])
        with col1:
            if st.button("🔄 Refresh", help="Refresh the document list"):
                st.rerun()
        
        # Filter controls
        col1, col2, col3 = st.columns(3)
        with col1:
            filter_ticker = st.text_input("Filter by Ticker", key="filter_ticker")
        with col2:
            filter_year = st.number_input("Filter by Year", min_value=2020, max_value=2030, value=None, key="filter_year")
        with col3:
            filter_quarter = st.selectbox("Filter by Quarter", [None, "Q1", "Q2", "Q3", "Q4"], key="filter_quarter")
        
        try:
            # Get uploaded documents
            docs = get_uploaded_documents(
                ticker=filter_ticker if filter_ticker else None,
                year=int(filter_year) if filter_year else None,
                quarter=filter_quarter if filter_quarter else None
            )
            
            if docs:
                # Show filtered results summary and clear button
                filter_summary = []
                if filter_ticker:
                    filter_summary.append(f"Ticker: {filter_ticker}")
                if filter_year:
                    filter_summary.append(f"Year: {filter_year}")
                if filter_quarter:
                    filter_summary.append(f"Quarter: {filter_quarter}")
                
                col1, col2 = st.columns([3, 1])
                with col1:
                    if filter_summary:
                        st.write(f"Found {len(docs)} documents matching: {', '.join(filter_summary)}")
                    else:
                        st.write(f"Found {len(docs)} uploaded documents:")
                
                with col2:
                    if filter_summary:
                        clear_button_text = f"🗑️ Clear Filtered ({len(docs)})"
                        clear_help_text = f"Delete {len(docs)} documents matching current filters"
                    else:
                        clear_button_text = f"🗑️ Clear All ({len(docs)})"
                        clear_help_text = f"Delete all {len(docs)} uploaded documents"
                    
                    if st.button(clear_button_text, help=clear_help_text, type="secondary"):
                        try:
                            # Delete each document in the filtered results
                            deleted_count = 0
                            for doc in docs:
                                try:
                                    # Call delete_document with both required parameters
                                    delete_document(doc.get('id'), doc.get('storage_path', ''))
                                    deleted_count += 1
                                except Exception as e:
                                    st.error(f"Failed to delete {doc.get('storage_path', 'unknown')}: {str(e)}")
                            
                            if deleted_count > 0:
                                st.success(f"✅ Deleted {deleted_count} documents successfully!")
                                st.rerun()
                            else:
                                st.warning("No documents were deleted")
                        except Exception as e:
                            st.error(f"Delete failed: {str(e)}")
                
                # Display documents with action buttons
                for i, doc in enumerate(docs):
                    with st.container():
                        col1, col2, col3, col4, col5, col6 = st.columns([2, 1, 1, 1, 1.2, 1])
                        
                        with col1:
                            # Extract filename from storage path
                            filename = doc.get('storage_path', '').split('/')[-1] if doc.get('storage_path') else 'Unknown'
                            st.write(f"**{filename}**")
                            st.caption(f"{doc.get('ticker', 'N/A')} • {doc.get('year', 'N/A')} • {doc.get('quarter', 'N/A')}")
                        
                        with col2:
                            st.write(doc.get('file_format', 'N/A').upper())
                        
                        with col3:
                            created_date = doc.get('created_at', '')
                            if created_date:
                                # Format date to show only date part
                                try:
                                    from datetime import datetime
                                    dt = datetime.fromisoformat(created_date.replace('Z', '+00:00'))
                                    st.write(dt.strftime('%Y-%m-%d'))
                                except:
                                    st.write(created_date[:10])
                            else:
                                st.write('N/A')
                        
                        with col4:
                            # File size if available
                            st.write('—')
                        
                        with col5:
                            # Download button
                            try:
                                # Download document content
                                doc_content = download_document(doc['storage_path'])
                                
                                # Direct download button
                                st.download_button(
                                    label="💾 Download",
                                    data=doc_content,
                                    file_name=filename,
                                    mime="application/octet-stream",
                                    key=f"download_{doc.get('id', i)}"
                                )
                            except Exception as e:
                                st.error(f"Failed to download: {str(e)}")
                        
                        with col6:
                            # Delete button
                            if st.button("🗑️ Delete", key=f"delete_{doc.get('id', i)}", help="Delete document", type="secondary"):
                                try:
                                    delete_document(doc['id'], doc['storage_path'])
                                    st.success(f"✅ Deleted '{filename}' successfully!")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"Failed to delete document: {str(e)}")
                    
                    st.divider()
            else:
                st.info("No uploaded documents found with current filters")
                
        except Exception as e:
            st.error(f"Error retrieving documents: {str(e)}")

# Move the processing logic back to the main guidance extraction tab
with main_tab1:
    # Robust CIK/ticker parsing
    user_input = ticker_or_cik.strip().upper()
    cik = None
    ticker = None

    if is_cik_format(user_input):
        cik = user_input
        ticker = get_ticker_from_cik(cik)
        if ticker:
            st.info(f"Using CIK {cik} for ticker {ticker}")
        else:
            st.info(f"Using CIK {cik} (ticker not found in SEC ticker file; will proceed with CIK)")
    elif user_input.isalnum():
        ticker = user_input
        cik = lookup_cik(ticker)
        if cik:
            st.info(f"Using ticker {ticker} (CIK: {cik})")
        else:
            st.error("CIK not found for ticker or input is not a valid CIK.")
            st.stop()
    else:
        st.error("Please enter a valid ticker (e.g., MSFT) or 10-digit CIK (e.g., 0000789019).")
        st.stop()

    if st.button("Extract Guidance", type="primary"):
        model_id = openai_models[selected_model]
        client = OpenAI(api_key=openai_key)
        st.info(f"Using OpenAI model: {selected_model}")
        
        all_results = []
        
        # SEC 8-K Processing
        if extract_sec:
            st.subheader("Processing SEC 8-K Filings")
            
            if quarter_input.strip():
                accessions = get_accessions(cik, ticker, specific_quarter=quarter_input.strip())
                if not accessions:
                    st.warning(f"No 8-K filings found for {quarter_input}. Please check the format (e.g., 2Q25, Q4FY24).")
            elif year_input.strip():
                try:
                    years_back = int(year_input.strip())
                    accessions = get_accessions(cik, ticker, years_back=years_back)
                except ValueError:
                    st.error("Invalid year input. Must be a number.")
                    accessions = []
            else:
                accessions = get_accessions(cik, ticker)

            if accessions:
                links = get_ex99_1_links(cik, accessions)
                
                for date_str, acc, url in links:
                    st.write(f"Processing {url}")
                    try:
                        headers = {"User-Agent": st.secrets.get('SEC_USER_AGENT', 'Your Name Contact@domain.com')}
                        html = requests.get(url, headers=headers, timeout=30).text
                        soup = BeautifulSoup(html, "html.parser")
                        text = soup.get_text(" ", strip=True)
                        guidance_paragraphs, found_guidance = find_guidance_paragraphs(text)

                        if found_guidance:
                            table = extract_guidance(guidance_paragraphs, ticker, client, model_id)
                            
                            df = process_guidance_table(table, "SEC 8-K", client, model_id)
                            if df is not None and not df.empty and len(df) > 0:
                                df["filing_date"] = date_str
                                df["filing_url"] = url
                                all_results.append(df)
                                st.success(f"Guidance extracted from this 8-K.")
                                st.dataframe(df[['metric', 'value_or_range', 'period', 'period_type']], use_container_width=True)
                            else:
                                st.warning(f"No guidance extracted from {url}")
                        else:
                            st.info(f"No guidance paragraphs found in this 8-K. Skipping.")
                            
                    except Exception as e:
                        st.error(f"Could not process: {url}. Error: {str(e)}")
        
        # Store SEC filing date for transcript use - capture even if no guidance found
        sec_filing_date = None
        if extract_sec and accessions:
            # Get filing date from processed 8-K filings (even if no guidance extracted)
            links = get_ex99_1_links(cik, accessions)
            if links:
                sec_filing_date = links[0][0]  # Use the most recent filing date
            
            # If we have results with guidance, use that date instead
            if all_results:
                for result in all_results:
                    if "filing_date" in result.columns and not result.empty:
                        sec_filing_date = result["filing_date"].iloc[0]
                        break

        # Transcript Processing
        if extract_transcripts:
            st.subheader("Processing Earnings Transcripts")
            
            # Parse quarter input for transcript search
            if quarter_input.strip():
                match = re.search(r'(?:Q?(\d)Q?|Q(\d))(?:\s*FY\s*|\s*)?(\d{2}|\d{4})', quarter_input.upper())
                if match:
                    quarter_num = int(match.group(1) or match.group(2))
                    year = match.group(3)
                    if len(year) == 2:
                        year = '20' + year
                    year_num = int(year)
                    
                    transcript, error, metadata = get_transcript_for_quarter(ticker, quarter_num, year_num)
                    if transcript:
                        # Extract guidance from transcript
                        st.write("Extracting guidance from transcript...")
                        table = extract_transcript_guidance(transcript, ticker, client, model_id)
                        df = process_guidance_table(table, "Transcript", client, model_id)
                        if df is not None and not df.empty:
                            # Use actual earnings date if available, otherwise fall back to SEC filing date or quarter format
                            earnings_date = metadata.get('earnings_date') if metadata else None
                            if earnings_date:
                                df["filing_date"] = earnings_date
                            elif sec_filing_date:
                                df["filing_date"] = sec_filing_date
                            else:
                                df["filing_date"] = f"{year_num}-Q{quarter_num}"
                            source = metadata.get('source', 'DefeatBeta') if metadata else 'DefeatBeta'
                            df["filing_url"] = f"{source} Transcript"
                            all_results.append(df)
                            st.success(f"Guidance extracted from transcript.")
                            st.dataframe(df[['metric', 'value_or_range', 'period', 'period_type']], use_container_width=True)
            elif year_input.strip():
                # Get transcripts for multiple quarters based on years back
                try:
                    years_back = int(year_input.strip())
                    current_year = datetime.now().year
                    
                    cutoff = datetime.now() - timedelta(days=(365 * years_back) + 91.25)
                    st.write(f"Looking for transcripts from the past {years_back} years plus 1 quarter (from {cutoff.strftime('%Y-%m-%d')} to present)")
                    
                    # Get fiscal year end information for proper quarter mapping
                    from edgar_enhanced import get_fiscal_year_end, get_fiscal_dates
                    fiscal_year_end_month, fiscal_year_end_day = get_fiscal_year_end(ticker, cik)
                    
                    # Get transcripts for the specified time period using fiscal year logic
                    current_date = datetime.now()
                    
                    # Calculate the most recent fiscal year based on current date and fiscal year end
                    if fiscal_year_end_month <= 3:  # Jan-Mar fiscal year end
                        if current_date.month <= fiscal_year_end_month:
                            current_fiscal_year = current_date.year
                        else:
                            current_fiscal_year = current_date.year + 1
                    else:  # Apr-Dec fiscal year end
                        if current_date.month > fiscal_year_end_month:
                            current_fiscal_year = current_date.year + 1
                        else:
                            current_fiscal_year = current_date.year
                    
                    # Track processed transcripts to avoid duplicates
                    processed_transcripts = set()
                    
                    # Get transcripts for fiscal years going back from current fiscal year
                    # Process the main years range first
                    for year_offset in range(years_back + 1):
                        target_fiscal_year = current_fiscal_year - year_offset
                        for quarter in [4, 3, 2, 1]:
                            # Get fiscal quarter information to determine proper calendar dates
                            fiscal_info = get_fiscal_dates(ticker, quarter, target_fiscal_year, fiscal_year_end_month, fiscal_year_end_day)
                            if fiscal_info:
                                # Use the fiscal year for transcript fetching
                                transcript, error, metadata = get_transcript_for_quarter(ticker, quarter, target_fiscal_year)
                            else:
                                transcript, error, metadata = get_transcript_for_quarter(ticker, quarter, target_fiscal_year)
                            if transcript:
                                # Use actual metadata for display instead of requested parameters
                                actual_quarter = metadata.get('quarter', f'Q{quarter}') if metadata else f'Q{quarter}'
                                actual_year = metadata.get('year', target_fiscal_year) if metadata else target_fiscal_year
                                
                                # Skip if requested period doesn't match actual metadata period
                                requested_quarter = f'Q{quarter}'
                                requested_year = target_fiscal_year
                                if actual_quarter != requested_quarter or actual_year != requested_year:
                                    st.info(f"Skipping {requested_quarter} {requested_year}: actual transcript is {actual_quarter} {actual_year}")
                                    continue
                                
                                # Create unique identifier for this transcript
                                transcript_id = f"{ticker}_{actual_quarter}_{actual_year}"
                                
                                # Skip if we've already processed this exact transcript
                                if transcript_id in processed_transcripts:
                                    st.info(f"Skipping duplicate: {actual_quarter} {actual_year} transcript already processed")
                                    continue
                                
                                processed_transcripts.add(transcript_id)
                                
                                st.write(f"Extracting guidance from {ticker} {actual_quarter} {actual_year} transcript...")
                                table = extract_transcript_guidance(transcript, ticker, client, model_id)
                                df = process_guidance_table(table, "Transcript", client, model_id)
                                if df is not None and not df.empty:
                                    # Use actual earnings date if available, otherwise fall back to quarter format using actual metadata
                                    earnings_date = metadata.get('earnings_date') if metadata else None
                                    if earnings_date:
                                        df["filing_date"] = earnings_date
                                    else:
                                        # Use actual metadata year/quarter, not the requested parameters
                                        df["filing_date"] = f"{actual_year}-{actual_quarter}"
                                    source = metadata.get('source', 'DefeatBeta') if metadata else 'DefeatBeta'
                                    df["filing_url"] = f"{source} Transcript {actual_quarter} {actual_year}"
                                    all_results.append(df)
                                    st.success(f"Guidance extracted from {actual_quarter} {actual_year} transcript.")
                                    st.dataframe(df[['metric', 'value_or_range', 'period', 'period_type']], use_container_width=True)
                    
                    # Process the extra quarter at the end (Q4 of the year before the range)
                    extra_year = current_fiscal_year - (years_back + 1)
                    transcript, error, metadata = get_transcript_for_quarter(ticker, 4, extra_year)
                    if transcript:
                        actual_quarter = metadata.get('quarter', 'Q4') if metadata else 'Q4'
                        actual_year = metadata.get('year', extra_year) if metadata else extra_year
                        
                        # Skip if requested period doesn't match actual metadata period
                        if actual_quarter == 'Q4' and actual_year == extra_year:
                            transcript_id = f"{ticker}_{actual_quarter}_{actual_year}"
                            if transcript_id not in processed_transcripts:
                                processed_transcripts.add(transcript_id)
                                st.write(f"Extracting guidance from {ticker} {actual_quarter} {actual_year} transcript...")
                                table = extract_transcript_guidance(transcript, ticker, client, model_id)
                                df = process_guidance_table(table, "Transcript", client, model_id)
                                if df is not None and not df.empty:
                                    earnings_date = metadata.get('earnings_date') if metadata else None
                                    report_date = metadata.get('report_date') if metadata else None
                                    
                                    if earnings_date and earnings_date not in [None, '', 'None']:
                                        df["filing_date"] = earnings_date
                                    elif report_date and report_date not in [None, '', 'None']:
                                        if hasattr(report_date, 'strftime'):
                                            df["filing_date"] = report_date.strftime('%Y-%m-%d')
                                        else:
                                            df["filing_date"] = str(report_date)
                                    else:
                                        df["filing_date"] = f"{actual_year}-{actual_quarter}"
                                    source = metadata.get('source', 'DefeatBeta') if metadata else 'DefeatBeta'
                                    df["filing_url"] = f"{source} Transcript {actual_quarter} {actual_year}"
                                    all_results.append(df)
                                    st.success(f"Guidance extracted from {actual_quarter} {actual_year} transcript.")
                                    st.dataframe(df[['metric', 'value_or_range', 'period', 'period_type']], use_container_width=True)
                                    
                except ValueError:
                    st.error("Invalid year input. Must be a number.")
            else:
                # Try to get most recent transcript
                transcript, error, metadata = get_transcript_for_quarter(ticker, None, None)
                if transcript:
                    st.write("Extracting guidance from most recent transcript...")
                    table = extract_transcript_guidance(transcript, ticker, client, model_id)
                    df = process_guidance_table(table, "Transcript", client, model_id)
                    if df is not None and not df.empty:
                        # Use actual earnings date if available, otherwise fall back to SEC filing date or "Most Recent"
                        earnings_date = metadata.get('earnings_date') if metadata else None
                        if earnings_date:
                            df["filing_date"] = earnings_date
                        elif sec_filing_date:
                            df["filing_date"] = sec_filing_date
                        else:
                            df["filing_date"] = "Most Recent"
                        source = metadata.get('source', 'DefeatBeta') if metadata else 'DefeatBeta'
                        df["filing_url"] = f"Most Recent {source} Transcript"
                        all_results.append(df)
                        st.success(f"Guidance extracted from transcript.")
                        st.dataframe(df[['metric', 'value_or_range', 'period', 'period_type']], use_container_width=True)
                    else:
                        st.warning("No guidance extracted from transcript.")

        # Uploaded Documents Processing
        if extract_uploaded:
            st.subheader("Processing Uploaded Documents")
            
            # Determine search parameters
            search_year = None
            search_quarter = None
            
            if quarter_input.strip():
                # Parse quarter input
                match = re.search(r'(?:Q?(\d)Q?|Q(\d))(?:\s*FY\s*|\s*)?(\d{2}|\d{4})', quarter_input.upper())
                if match:
                    quarter_num = int(match.group(1) or match.group(2))
                    year = match.group(3)
                    if len(year) == 2:
                        year = '20' + year
                    search_year = int(year)
                    search_quarter = f"Q{quarter_num}"
            
            try:
                # Get uploaded documents for this ticker/year/quarter
                uploaded_docs = get_uploaded_documents(
                    ticker=ticker,
                    year=search_year,
                    quarter=search_quarter
                )
                
                if uploaded_docs:
                    st.write(f"Found {len(uploaded_docs)} uploaded documents for {ticker}")
                    
                    for doc in uploaded_docs:
                        st.write(f"Processing: {doc['file_type']} - {doc.get('storage_path', 'Unknown path')}")
                        
                        try:
                            # Download document content using the enhanced bucket management
                            doc_content = download_document(doc['storage_path'])
                            
                            # Extract text content based on file format
                            text_content = None
                            if doc['file_format'] in ['txt', 'html']:
                                text_content = doc_content.decode('utf-8')
                            elif doc.get('text_content'):
                                text_content = doc['text_content']
                            elif doc['file_format'] in ['pdf', 'docx']:
                                # For PDF/DOCX files, try to extract text from the downloaded content
                                try:
                                    from document_tagger import extract_text_from_file
                                    text_content = extract_text_from_file(doc_content, doc['file_format'])
                                except Exception as extract_error:
                                    st.warning(f"Could not extract text from {doc['file_format']}: {str(extract_error)}")
                                    continue
                            else:
                                st.warning(f"Cannot extract text from {doc['file_format']} format yet")
                                continue
                            
                            if text_content:
                                # Smart content filtering for large documents
                                st.info(f"Document size: {len(text_content):,} characters")
                                
                                # Use efficient guidance paragraph extraction first
                                guidance_paragraphs, found_guidance = find_guidance_paragraphs(text_content)
                                
                                if found_guidance:
                                    st.success(f"Found potential guidance in {doc['file_type']} ({len(guidance_paragraphs):,} chars after filtering)")
                                    table = extract_guidance(guidance_paragraphs, ticker, client, model_id)
                                    
                                    df = process_guidance_table(table, f"Uploaded {doc['file_type']}", client, model_id)
                                    if df is not None and not df.empty:
                                        df["filing_date"] = f"{doc['year']}-{doc['quarter']}"
                                        df["filing_url"] = f"Uploaded: {doc.get('storage_path', 'Unknown')}"
                                        all_results.append(df)
                                        st.success(f"Guidance extracted from uploaded {doc['file_type']}")
                                        st.dataframe(df[['metric', 'value_or_range', 'period', 'period_type']], use_container_width=True)
                                    else:
                                        st.warning(f"No guidance extracted from {doc['file_type']}")
                                else:
                                    st.info(f"No guidance paragraphs found in {doc['file_type']}")
                                    
                        except Exception as e:
                            st.error(f"Error processing {doc['file_type']}: {str(e)}")
                            
                else:
                    st.info(f"No uploaded documents found for {ticker}" + 
                           (f" {search_year}-{search_quarter}" if search_year and search_quarter else ""))
                    
            except Exception as e:
                st.error(f"Error retrieving uploaded documents: {str(e)}")

        # Display combined results
        if all_results:
            combined = pd.concat(all_results, ignore_index=True)
            
            # Display rename for better presentation
            display_rename = {
                'metric': 'Metric',
                'value_or_range': 'Value or Range',
                'period': 'Period',
                'period_type': 'Period Type',
                'low': 'Low',
                'high': 'High',
                'average': 'Average',
                'filing_date': 'Filing Date',
                'filing_url': 'Source'
            }
            
            # Preview table
            st.subheader("Preview of Extracted Guidance")
            display_cols = ['metric', 'value_or_range', 'period', 'period_type', 'low', 'average', 'high', 'filing_date', 'source_type']
            display_df = combined[display_cols] if all(col in combined.columns for col in display_cols) else combined
            display_df = display_df.rename(columns={c: display_rename.get(c, c) for c in display_df.columns})
            st.dataframe(display_df, use_container_width=True)
            
            # Check for duplicates and highlight them
            duplicate_indices = detect_duplicates(combined, client, model_id)
            
            if duplicate_indices:
                st.warning(f"⚠️ Found {len(duplicate_indices)} duplicate guidance entries highlighted in yellow.")
                
                # Display dataframe with duplicate highlighting
                styled_df = highlight_duplicates(combined, duplicate_indices)
                st.dataframe(styled_df, use_container_width=True)
            
            # Always show download - no duplicate resolution needed
            st.success("✅ Results ready for download!")
            
            # Create Excel download with duplicate highlighting
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                # Reorder columns to flip filing_date and source_type
                column_order = list(combined.columns)
                if 'filing_date' in column_order and 'source_type' in column_order:
                    filing_date_idx = column_order.index('filing_date')
                    source_type_idx = column_order.index('source_type')
                    column_order[filing_date_idx], column_order[source_type_idx] = column_order[source_type_idx], column_order[filing_date_idx]
                
                combined_reordered = combined[column_order]
                combined_reordered.to_excel(writer, sheet_name='Guidance_Data', index=False)
                
                # Apply formatting to Excel
                from openpyxl.styles import PatternFill, Alignment
                worksheet = writer.sheets['Guidance_Data']
                
                # Center align all columns except metric (column A)
                center_alignment = Alignment(horizontal='center')
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=2, max_col=worksheet.max_column):
                    for cell in row:
                        cell.alignment = center_alignment
                
                # Apply highlighting to duplicates in Excel
                if duplicate_indices:
                    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    
                    for idx in duplicate_indices:
                        # Excel rows are 1-indexed and we need to account for header
                        excel_row = idx + 2
                        for col in range(1, len(combined.columns) + 1):
                            worksheet.cell(row=excel_row, column=col).fill = yellow_fill
            
            excel_buffer.seek(0)  # Reset buffer position
            
            st.download_button(
                "📥 Download Excel",
                data=excel_buffer.getvalue(),
                file_name=f"{ticker}_guidance_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("No guidance data extracted from any source.")
            st.info("Try adjusting your search criteria or check if the company provides guidance in their earnings materials.")

# Legacy code section - keeping for compatibility
if False:  # Disabled duplicate section
    st.subheader("Combined Guidance Results")
    combined = pd.concat(all_results, ignore_index=True)
    
    # Display rename for better presentation
    display_rename = {
        'metric': 'Metric',
        'value_or_range': 'Value or Range',
        'period': 'Period',
        'period_type': 'Period Type',
        'low': 'Low',
        'high': 'High',
        'average': 'Average',
        'filing_date': 'Filing Date',
        'filing_url': 'Source'
    }
    
    # Preview table
    st.subheader("Preview of Extracted Guidance")
    display_cols = ['metric', 'value_or_range', 'period', 'period_type', 'low', 'average', 'high', 'filing_date', 'source_type']
    display_df = combined[display_cols] if all(col in combined.columns for col in display_cols) else combined
    display_df = display_df.rename(columns={c: display_rename.get(c, c) for c in display_df.columns})
    st.dataframe(display_df, use_container_width=True)
    
    # Check for duplicates and highlight them
    duplicate_indices = detect_duplicates(combined, client, model_id)
    
    if duplicate_indices:
        st.warning(f"⚠️ Found {len(duplicate_indices)} duplicate guidance entries highlighted in yellow.")
        
        # Display dataframe with duplicate highlighting
        styled_df = highlight_duplicates(combined, duplicate_indices)
        st.dataframe(styled_df, use_container_width=True)
    
    # Always show download - no duplicate resolution needed
    st.success("✅ Results ready for download!")
    
    # Create Excel download with duplicate highlighting
    excel_buffer = BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        # Reorder columns to flip filing_date and source_type
        column_order = list(combined.columns)
        if 'filing_date' in column_order and 'source_type' in column_order:
            filing_date_idx = column_order.index('filing_date')
            source_type_idx = column_order.index('source_type')
            column_order[filing_date_idx], column_order[source_type_idx] = column_order[source_type_idx], column_order[filing_date_idx]
        
        combined_reordered = combined[column_order]
        combined_reordered.to_excel(writer, sheet_name='Guidance_Data', index=False)
        
        # Apply formatting to Excel
        from openpyxl.styles import PatternFill, Alignment
        worksheet = writer.sheets['Guidance_Data']
        
        # Center align all columns except metric (column A)
        center_alignment = Alignment(horizontal='center')
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=2, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = center_alignment
        
        # Apply highlighting to duplicates in Excel
        if duplicate_indices:
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            
            for idx in duplicate_indices:
                # Excel rows are 1-indexed and we need to account for header
                excel_row = idx + 2
                for col in range(1, len(combined.columns) + 1):
                    worksheet.cell(row=excel_row, column=col).fill = yellow_fill
    
    excel_buffer.seek(0)  # Reset buffer position
    
    st.download_button(
        "📥 Download Excel",
        data=excel_buffer.getvalue(),
        file_name="uploaded_guidance_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

st.markdown("---")
st.markdown("**Tips:**")
st.markdown("- Use specific quarters (e.g., 'Q1 2024') for targeted searches")
st.markdown("- Check SEC filings, transcripts, and any user uploaded documents (ex. earning presentations) for comprehensive guidance")
st.markdown("- The Excel download includes all extracted data with source attribution")

